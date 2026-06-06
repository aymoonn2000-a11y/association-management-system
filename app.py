import streamlit as st
import pandas as pd
import json
import os
import hashlib
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- 1. إعدادات الصفحة والديزاين العام المتجاوب مع الجوال والكمبيوتر ---
st.set_page_config(
    page_title="نظام إدارة الجمعية المتكامل",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# دالة لتشفير كلمة المرور لزيادة الأمان حماية البيانات
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# تذكر تغيير كلمة المرور هنا أو التحقق منها مشفرة
# كلمة المرور الافتراضية "12345" مشفرة بـ SHA-256 هي المكتوبة أدناه
CORRECT_USERNAME = "aymanyaghi"
CORRECT_PASSWORD_HASH = "5994471abb01112afcc18159f6cc74b4f511b99806da59b3caf5a9c173cacfc5" 

# ستايل CSS متطور لتحسين المظهر على الجوال ولاب توب وأيقونات منسقة
st.markdown("""
<style>
.main { text-align: right; direction: rtl; }
[data-testid="stSidebar"] { direction: rtl; }
div.stButton > button:first-child {
    background-color: #1f77b4;
    color: white;
    border-radius: 8px;
    padding: 0.6rem 2rem;
    font-size: 16px;
    width: 100%;
    font-weight: bold;
}
h1, h2, h3, p, span, div {
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif !important;
    text-align: right;
}

/* شعار المنظومة الاحترافي */
.logo-container {
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 12px;
    margin-bottom: 20px;
    background: #ffffff;
    padding: 12px;
    border-radius: 12px;
    box-shadow: 0 4px 10px rgba(0,0,0,0.05);
}
.logo-icon {
    width: 45px;
    height: 45px;
    background: linear-gradient(135deg, #1f77b4, #00d2ff);
    color: white;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 22px;
    font-weight: bold;
}
.logo-text {
    font-size: 22px;
    font-weight: 800;
    color: #1f77b4;
    font-family: 'Century Gothic', sans-serif;
}

/* بطاقة الترحيب */
.welcome-card {
    background: linear-gradient(135deg, #1f77b4, #00d2ff);
    color: white;
    padding: 20px;
    border-radius: 15px;
    text-align: center;
    margin-bottom: 25px;
    box-shadow: 0 4px 15px rgba(0,0,0,0.1);
}
.info-box {
    background-color: #ffffff;
    padding: 15px;
    border-radius: 12px;
    border-right: 5px solid #1f77b4;
    box-shadow: 0 4px 6px rgba(0,0,0,0.05);
    text-align: center;
}
</style>
""", unsafe_allow_html=True)

# --- 2. دالات تخزين وحفظ البيانات الدائمة محلياً (بدون انترنت JSON) ---
def load_data(filename):
    if os.path.exists(filename):
        with open(filename, 'r', encoding='utf-8') as f:
            try:
                return json.load(f)
            except:
                return []
    return []

def save_data(filename, data):
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

# دالة تصدير ملف إكسل منسق، ملون، مفلتر ومطابق لنموذج الجمعية المرفق
def export_to_styled_excel(dataframe, title_report="تقرير عام", is_transport=False):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "البيانات"
    ws.views.sheetView[0].rightToLeft = True # تفعيل من اليمين لليسار للغة العربية
    
    # تنسيقات الألوان والخطوط (الاستايل الملون والفلترة)
    header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
    header_font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=11)
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1F77B4")
    accent_fill = PatternFill(start_color="E6F2FA", end_color="E6F2FA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    current_row = 1
    
    # إذا كان كشف مواصلات مثل الملف المرفق نضيف ترويسة الجمعية الرسمية
    if is_transport:
        ws.cell(row=1, column=2, value="جمعية الحياة والأمل").font = Font(name="Segoe UI", size=12, bold=True)
        ws.cell(row=1, column=7, value="Life and Hope Association").font = Font(name="Segoe UI", size=12, bold=True)
        ws.cell(row=3, column=2, value="محافظة شمال غزة").font = Font(name="Segoe UI", size=11, italic=True)
        ws.cell(row=3, column=7, value="Gaza – North Governorate").font = Font(name="Segoe UI", size=11, italic=True)
        
        ws.merge_cells("A5:I5")
        title_cell = ws.cell(row=5, column=1, value=f"كشف مواصلات - {title_report}")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center")
        current_row = 7
    else:
        ws.cell(row=1, column=1, value=title_report).font = title_font
        current_row = 3
        
    # كتابة أسماء الأعمدة (الترويسة)
    columns = list(dataframe.columns)
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=current_row, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    start_data_row = current_row + 1
    current_row += 1
    
    # كتابة البيانات الصفوف وتلوينها بالتناوب
    for index, row in dataframe.iterrows():
        for col_num, cell_value in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=col_num, value=row[cell_value])
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if index % 2 == 0:
                cell.fill = accent_fill
        current_row += 1
        
    end_data_row = current_row - 1
    
    # إضافة الفلترة التلقائية للإكسل لجميع الأعمدة المحددة
    ws.auto_filter.ref = f"A{start_data_row-1}:{get_column_letter(len(columns))}{end_data_row}"
    
    # إضافة خلايا الدوال الحسابية والتوقيعات لملف المواصلات المرفق
    if is_transport and "المبلغ بالشيكل" in dataframe.columns:
        amount_col_idx = dataframe.columns.get_loc("المبلغ بالشيكل") + 1
        current_row += 1
        ws.cell(row=current_row, column=amount_col_idx-1, value="المبلغ الإجمالي بالشيكل:").font = Font(name="Segoe UI", size=11, bold=True)
        # دالة الجمع التلقائي بالإكسل SUM
        sum_cell = ws.cell(row=current_row, column=amount_col_idx, value=f"=SUM({get_column_letter(amount_col_idx)}{start_data_row}:{get_column_letter(amount_col_idx)}{end_data_row})")
        sum_cell.font = Font(name="Segoe UI", size=11, bold=True)
        sum_cell.border = Border(bottom=Side(style='double', color='000000'), top=Side(style='thin', color='000000'))
        
        # تذييل التوقيعات الرسمية كما بالملف الاصلي للجمعية
        current_row += 3
        signatures = [
            ("رئيس مجلس الإدارة", "زاهر محمود صبيح"),
            ("مدير/منسق المشروع", "........................"),
            ("المدير المالي", "........................")
        ]
        col_positions = [2, 5, 8]
        for pos, (title, name) in zip(col_positions, signatures):
            ws.cell(row=current_row, column=pos, value=f"المسمى الوظيفي / {title}").font = Font(name="Segoe UI", size=10, bold=True)
            ws.cell(row=current_row+1, column=pos, value=f"الاسم / {name}").font = Font(name="Segoe UI", size=10)
            ws.cell(row=current_row+2, column=pos, value="التوقيع / ....................................").font = Font(name="Segoe UI", size=10)

    # ضبط العرض التلقائي للأعمدة لتجنب التفاف النص المشوه
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output)
    return output.getvalue()

# --- 3. تهيئة الجلسات وتحميل البيانات التلقائي المخزن محلياً ---
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
if 'expenses' not in st.session_state:
    st.session_state.expenses = load_data('expenses.json')
if 'employees' not in st.session_state:
    st.session_state.employees = load_data('employees.json')
if 'inventory' not in st.session_state:
    st.session_state.inventory = load_data('inventory.json')
if 'transport_records' not in st.session_state:
    st.session_state.transport_records = load_data('transport_records.json')

# --- 4. نظام تسجيل الدخول المحمي والآمن ---
if not st.session_state.logged_in:
    st.markdown("<h1 style='text-align: center;'>🔒 تسجيل الدخول الآمن لنظام الجمعية</h1>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        username = st.text_input("👤 اسم المستخدم", placeholder="أدخل اسم المستخدم هنا")
        password = st.text_input("🔑 كلمة المرور", type="password", placeholder="أدخل كلمة المرور هنا")
        login_button = st.button("🚪 دخول نظام الإدارة")
        
        if login_button:
            # مقارنة كلمة المرور المدخلة بعد تحويلها الـ Hash لحماية أمنية مشددة
            if username == CORRECT_USERNAME and hash_password(password) == CORRECT_PASSWORD_HASH:
                st.session_state.logged_in = True
                st.success("🔓 تم التحقق وتأمين الجلسة بنجاح!")
                st.rerun()
            else:
                st.error("❌ اسم المستخدم أو كلمة المرور غير مطابقة لمقاييس الحماية!")

# --- 5. لوحة التحكم بعد تسجيل الدخول الآمن ---
else:
    # عرض الهوية البرمجية في شريط جانبي
    st.sidebar.markdown("""
    <div class="logo-container">
        <div class="logo-icon">🏢</div>
        <div>
            <div class="logo-text">جمعية الحياة والأمل</div>
            <div style="font-size:12px; color:#555; text-align:right;">نظام تتبع لوجستي متكامل</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    st.sidebar.markdown("<h2 style='text-align: center; color: #1f77b4;'>📋 شاشة التنقل الذكية</h2>", unsafe_allow_html=True)
    
    menu = st.sidebar.radio(
        "اختر القسم البرمجي العملياتي:",
        [
            "🏠 الشاشة الرئيسية والأداء", 
            "🚗 كشف مواصلات فردي (الملف المرفق)", 
            "💰 المصروفات اليومية والشهرية", 
            "👥 شؤون الموظفين والعهدة", 
            "📦 جرد مواد المخزن والمكتب", 
            "🚪 تسجيل الخروج الآمن"
        ]
    )
    
    if menu == "🚪 تسجيل الخروج الآمن":
        st.session_state.logged_in = False
        st.rerun()

    # --- القسم 1: الشاشة الرئيسية ولوحة التحليلات والبيانات الجغرافية والزمنية ---
    elif menu == "🏠 الشاشة الرئيسية والأداء":
        st.markdown("""
        <div class="welcome-card">
            <h1 style="color: white; margin: 0; font-family: 'Century Gothic', sans-serif; text-align:center;">مرحباً بك في لوحة القيادة المتكاملة</h1>
            <p style="font-size: 18px; margin-top: 10px; text-align:center;">أداء النظام والتحليلات الإحصائية الفورية - متوافق مع أجهزة الهاتف والحاسوب الشخصي</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"""
            <div class="info-box">
                <h3 style="margin:0; color:#1f77b4;">📅 تاريخ الخادم المحلي</h3>
                <p style="font-size: 20px; font-weight: bold; margin-top: 10px;">{datetime.now().strftime('%A, %Y-%m-%d')}</p>
            </div>
            """, unsafe_allow_html=True)
            
        with col2:
            st.markdown(f"""
            <div class="info-box">
                <h3 style="margin:0; color:#1f77b4;">⏰ الوقت الفعلي الحالي</h3>
                <p style="font-size: 20px; font-weight: bold; margin-top: 10px;">{datetime.now().strftime('%I:%M %p')}</p>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br><h3>📊 الهيكل البياني التوزيعي للمصروفات والبنود المخصصة</h3>", unsafe_allow_html=True)
        if st.session_state.expenses:
            df_all = pd.DataFrame(st.session_state.expenses)
            df_chart = df_all.groupby("نوع السلعة")["السعر (شيكل)"].sum().reset_index()
            st.bar_chart(data=df_chart, x="نوع السلعة", y="السعر (شيكل)", use_container_width=True)
        else:
            st.info("💡 لا يوجد حركة قيود مالية كافية لتوليد رسم بياني حالياً.")

    # --- القسم 2: كشف مواصلات فردي (تطبيق ملف الإكسل المرفق بالكامل وبنفس البنية والدوال) ---
    elif menu == "🚗 كشف مواصلات فردي (الملف المرفق)":
        st.title("🚗 نموذج كشف مواصلات الموظفين الفردي")
        st.info("📝 هذا القسم تمت صياغته وهندسته برمجياً ليطابق ملف الإكسل المرفق الخاص بـ (جمعية الحياة والأمل - محافظة شمال غزة)")
        st.markdown("---")
        
        col1, col2 = st.columns([1, 2])
        with col1:
            st.subheader("➕ تسجيل حركة مواصلات جديدة")
            trans_emp = st.text_input("👤 اسم الموظف الحالي", key="trans_emp")
            trans_job = st.text_input("💼 المسمى الوظيفي للموظف", key="trans_job")
            trans_date = st.date_input("📅 تاريخ الحركة اليومية", datetime.now())
            trans_day = st.selectbox("📆 اليوم", ["السبت", "الأحد", "الإثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة"])
            trans_from = st.text_input("📍 نقطة الانطلاق (من)")
            trans_to = st.text_input("🏁 نقطة الوصول (إلى)")
            trans_amount = st.number_input("💰 المبلغ المستحق بالشيكل", min_value=0.0, step=0.5)
            trans_reason = st.text_area("🔍 سبب الحركة / طبيعة المهمة الميدانية")
            
            if st.button("💾 ترحيل وحفظ الحركة الميدانية"):
                if trans_emp and trans_from and trans_to and trans_amount > 0:
                    st.session_state.transport_records.append({
                        "id": datetime.now().strftime("%Y%m%d%H%M%S"),
                        "اسم الموظف": trans_emp,
                        "المسمى الوظيفي": trans_job,
                        "اليوم": trans_day,
                        "التاريخ": trans_date.strftime("%Y-%m-%d"),
                        "من": trans_from,
                        "إلى": trans_to,
                        "المبلغ بالشيكل": trans_amount,
                        "سبب الحركة": trans_reason
                    })
                    save_data('transport_records.json', st.session_state.transport_records)
                    st.success("✅ تم حفظ قيد حركة المواصلات بنجاح في قاعدة البيانات المحلية!")
                    st.rerun()
                else:
                    st.error("⚠️ يرجى تعبئة كافة الحقول الأساسية وتحديد قيمة مالية صحيحة!")
                    
        with col2:
            st.subheader("🔍 استعراض الكشوفات الميدانية وتصدير ملف Excel الملون")
            if st.session_state.transport_records:
                df_trans = pd.DataFrame(st.session_state.transport_records)
                
                # تصفية وفلترة البيانات حسب اسم الموظف لتسهيل المتابعة الفردية
                unique_emps = df_trans["اسم الموظف"].unique()
                filter_emp = st.selectbox("🎯 تصفية عرض البيانات بحسب الموظف:", unique_emps)
                
                df_display = df_trans[df_trans["اسم الموظف"] == filter_emp].copy()
                job_title_display = df_display["المسمى الوظيفي"].iloc[0] if not df_display.empty else ""
                
                st.write(f"**🤵 موظف الكشف:** {filter_emp} | **📋 المسمى الوظيفي:** {job_title_display}")
                
                # إظهار الجدول منسق ومحدد
                st.dataframe(df_display.drop(columns=["id", "اسم الموظف", "المسمى الوظيفي"]), use_container_width=True)
                
                # حساب المجموع الفوري على الشاشة
                total_sum = df_display["المبلغ بالشيكل"].sum()
                st.metric(label="📊 إجمالي المبالغ المستحقة (شيكل)", value=f"{total_sum} ₪")
                
                # زر التنزيل لملف Excel الملون والمنسق بالكامل والمحتوي على الفلاتر التلقائية والدوال الحسابية والتوقيعات
                excel_data = export_to_styled_excel(
                    df_display.drop(columns=["id"]), 
                    title_report=f"الموظف {filter_emp}", 
                    is_transport=True
                )
                
                st.download_button(
                    label="📥 تنزيل كشف المواصلات بصيغة Excel (منسق وملون ومفلتر تلقائياً)",
                    data=excel_data,
                    file_name=f"كشف_مواصلات_{filter_emp}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # حذف حركة معينة من الكشف
                st.markdown("---")
                delete_trans_id = st.selectbox("🗑️ إلغاء أو حذف حركة من القائمة (حسب الوجهة والسبب):", df_display["id"].tolist(), format_func=lambda x: f"من {next(i['من'] for i in st.session_state.transport_records if i['id'] == x)} إلى {next(i['إلى'] for i in st.session_state.transport_records if i['id'] == x)} - {next(i['المبلغ بالشيكل'] for i in st.session_state.transport_records if i['id'] == x)} شيكل")
                if st.button("❌ حذف الحركة المحددة"):
                    st.session_state.transport_records = [i for i in st.session_state.transport_records if i["id"] != delete_trans_id]
                    save_data('transport_records.json', st.session_state.transport_records)
                    st.success("🗑️ تم حذف القيد اللوجستي من السجلات المحلية!")
                    st.rerun()
            else:
                st.info("💡 لا يوجد حركات مواصلات مدخلة حتى اللحظة.")

    # --- القسم 3: المصروفات العمومية ---
    elif menu == "💰 المصروفات اليومية والشهرية":
        st.title("💰 إدارة البنود المالية والمصروفات العمومية")
        st.markdown("---")
        
        col1, col2 = st.columns([1, 2])
        with col1:
            st.subheader("➕ قيد بند مالي جديد")
            exp_date = st.date_input("تاريخ الصرف المباشر", datetime.now())
            exp_types = st.multiselect("🏷️ نوع السلعة والتبويب المالي", ["مواصلات تابعة للمقر", "ضيافة ومياه", "قرطاسية وأوراق", "صيانة مكتبية", "أخرى"], default=["ضيافة ومياه"])
            exp_details = st.text_input("📑 الوصف التفصيلي للفاتورة / السند")
            exp_price = st.number_input("💸 القيمة المالية (شيكل)", min_value=0.0, step=1.0)
            
            if st.button("💾 ترحيل البند المالي"):
                if exp_types and exp_price > 0:
                    types_str = ", ".join(exp_types)
                    st.session_state.expenses.append({
                        "id": datetime.now().strftime("%Y%m%d%H%M%S"),
                        "التاريخ": exp_date.strftime("%Y-%m-%d"),
                        "الشهر المستهدف": exp_date.strftime("%Y-%m"),
                        "نوع السلعة": types_str,
                        "التفاصيل": exp_details,
                        "السعر (شيكل)": exp_price
                    })
                    save_data('expenses.json', st.session_state.expenses)
                    st.success("✅ تم قيد التكلفة وحفظها محلياً!")
                    st.rerun()
                else:
                    st.warning("⚠️ يرجى التأكد من ملء البيانات الأساسية وقيمة الفاتورة!")
                    
        with col2:
            st.subheader("📊 كشوفات تصفية الحسابات والتحميل الملون")
            if st.session_state.expenses:
                df_all = pd.DataFrame(st.session_state.expenses)
                available_months = sorted(list(df_all["الشهر المستهدف"].unique()), reverse=True)
                selected_month = st.selectbox("📂 اختيار دورة التصفية (الشهر المستهدف):", available_months)
                
                df_filtered = df_all[df_all["الشهر المستهدف"] == selected_month].copy()
                st.dataframe(df_filtered.drop(columns=["id", "الشهر المستهدف"]), use_container_width=True)
                
                # زر تنزيل Excel الملون للمصروفات العمومية
                excel_exp = export_to_styled_excel(df_filtered.drop(columns=["id"]), title_report=f"تقرير مصروفات شهر {selected_month}", is_transport=False)
                st.download_button(
                    label="📥 تحميل كشف المصروفات منسق وملون ومليء بالفلاتر التلقائية (Excel)",
                    data=excel_exp,
                    file_name=f"مصروفات_{selected_month}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.markdown("---")
                delete_id = st.selectbox("🗑️ اختيار بند لإسقاطه وحذفه من السجلات المالي:", df_filtered["id"].tolist(), format_func=lambda x: next(item["التفاصيل"] for item in st.session_state.expenses if item["id"] == x))
                if st.button("❌ حذف البند المالي المحدد"):
                    st.session_state.expenses = [item for item in st.session_state.expenses if item["id"] != delete_id]
                    save_data('expenses.json', st.session_state.expenses)
                    st.success("🗑️ تم حذف السند ومزامنة البيانات الدائمة.")
                    st.rerun()

    # --- القسم 4: شؤون الموظفين والعهدة المستلمة والمستردة ---
    elif menu == "👥 شؤون الموظفين والعهدة":
        st.title("👥 إدارة الموظفين وسجلات العهدة والتتبع اللوجستي")
        st.markdown("---")
        
        col1, col2 = st.columns([1, 2])
        with col1:
            st.subheader("➕ تسجيل بيانات موظف")
            emp_name = st.text_input("🤵 الاسم ثلاثي / رباعي")
            emp_id = st.text_input("💳 رقم الهوية الوطنية / جواز السفر")
            emp_phone = st.text_input("📱 رقم الهاتف النقال للتواصل")
            emp_title = st.text_input("💼 المسمى أو التوصيف الوظيفي داخل الهيكل")
            emp_program = st.selectbox("🏢 البرنامج / القسم التابع له العمل الميداني", ["Shelter", "CVA", "WASH", "Program", "Administrative Assistant", "Manager"])
            emp_assets = st.multiselect("📦 العهد والأصول المستلمة للتأدية الميدانية", ["تيشيرت الجمعية", "لاب توب", "أيباد / تابلت", "قرطاسية كاملة", "أخرى"])
            asset_status = st.selectbox("🔄 وضع وحالة العهدة الحالية", ["مستلمة بالكامل وفي حوزة الموظف", "تم تصفيتها وإرجاعها للمخزن المركزي"])
            return_date
